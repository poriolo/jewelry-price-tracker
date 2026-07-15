"""
Gera modelos de dívida em Excel — módulos acopláveis a um valuation.
Config-driven: cada empresa é um dicionário (períodos, tranches, âncoras reais).
Periodicidade: trimestral (aberto até 4T27) e depois anual até 2037.
Células são fórmulas vivas: amarelo=input, azul=real, verde=interface.

Rode:  python3 build_debt_model.py
Saída: vivara_debt_model.xlsx e smartfit_debt_model.xlsx

Fontes dos dados-âncora estão na aba "Leia-me" de cada arquivo e no README.md.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ----------------------------------------------------------------------------
# Curva de CDI compartilhada (premissa, % a.a.) — trimestral e depois anual
# ----------------------------------------------------------------------------
Q = [  # (label, fator, cdi)
    ("2T26E", 0.25, 0.1490), ("3T26E", 0.25, 0.1475), ("4T26E", 0.25, 0.1450),
    ("1T27E", 0.25, 0.1400), ("2T27E", 0.25, 0.1350), ("3T27E", 0.25, 0.1300),
    ("4T27E", 0.25, 0.1250),
]
A = [  # anos (label, fator, cdi)
    ("2028E", 1.0, 0.1100), ("2029E", 1.0, 0.1050), ("2030E", 1.0, 0.1000),
    ("2031E", 1.0, 0.0975), ("2032E", 1.0, 0.0950), ("2033E", 1.0, 0.0950),
    ("2034E", 1.0, 0.0950), ("2035E", 1.0, 0.0950), ("2036E", 1.0, 0.0950),
    ("2037E", 1.0, 0.0950),
]

# ----------------------------------------------------------------------------
# Configurações das empresas
# ----------------------------------------------------------------------------
VIVARA = {
    "file": "vivara_debt_model.xlsx",
    "title": "VIVARA (VIVA3) — Fluxo de Dívida (módulo de valuation)",
    "periods": (
        [("4T25\n(Real)", "anchor", None, None), ("1T26\n(Real)", "real", 0.25, 0.1490)]
        + [(l, "proj", f, c) for l, f, c in Q] + [(l, "proj", f, c) for l, f, c in A]
    ),
    "ir": 0.34,
    "tranches": [
        {"name": "Debêntures VIVA11", "desc": "100% CDI + 0,70% • amort. 50% 2029 / 50% 2030",
         "opening": 300.0, "spread": 0.0070, "amort": {"2029E": 150.0, "2030E": 150.0}},
        {"name": "Outras dívidas", "desc": "bancos / FX / capital de giro — premissa: rolagem",
         "opening": 231.3, "spread": 0.0200, "amort": {}},
    ],
    "cash_anchors": {"4T25\n(Real)": 398.6, "1T26\n(Real)": 284.7},
    "readme": [
        ("PREMISSAS-CHAVE (reais, pesquisadas)", "b"),
        ("• Debênture VIVA11 (1ª emissão): R$ 300 mi, 27/08/2025→2030, 100% CDI + 0,70% a.a.,", "n"),
        ("  juros semestrais; amortização 50% em 2029 e 50% em 2030; recursos p/ liquidar CCBs.", "n"),
        ("• 31/12/2025: dívida bruta R$ 531,3 mi | caixa R$ 398,6 mi | dívida líquida R$ 132,6 mi.", "n"),
        ("• 1T26: dívida líquida R$ 246,6 mi | alavancagem 0,3x ND/EBITDA.", "n"),
        ("", None),
        ("PREMISSAS ESTIMADAS (ajuste com o ITR)", "b"),
        ("• 'Outras dívidas' = bruta − debêntures = R$ 231,3 mi, modeladas como rolagem (amort. = 0).", "n"),
        ("• Spread outras dívidas = CDI + 2,00% a.a.; caixa 1T26 (R$ 284,7 mi) fecha a DL reportada.", "n"),
        ("• 'Despesa de juros' é só a da dívida; o resultado financeiro reportado (−R$ 15 mi no 1T26)", "n"),
        ("  inclui derivativos/MtM e não bate 1:1.", "n"),
        ("", None),
        ("FONTES", "b"),
        ("• Release 1T26 (Money Times / Nord / Análise de Ações); VIVA11 (debentures.com.br /", "s"),
        ("  Mais Retorno / ANBIMA); emissão R$ 300 mi (Lex Legal); DF 2025 Vivara Participações.", "s"),
    ],
}

SMARTFIT = {
    "file": "smartfit_debt_model.xlsx",
    "title": "SMART FIT (SMFT3) — Fluxo de Dívida (módulo de valuation)",
    "periods": (
        [("1T26\n(Real)", "anchor", None, None)]
        + [(l, "proj", f, c) for l, f, c in Q] + [(l, "proj", f, c) for l, f, c in A]
    ),
    "ir": 0.34,
    "tranches": [
        {"name": "Dívida financeira", "desc": "custo médio CDI + 2,09% • premissa: rolagem",
         "opening": 5900.0, "spread": 0.0209, "amort": {}},
    ],
    "cash_anchors": {"1T26\n(Real)": 1700.0},
    "readme": [
        ("PREMISSAS-CHAVE (reais, pesquisadas — 1T26)", "b"),
        ("• Dívida líquida R$ 4,2 bi (estável vs 4T25) | caixa+títulos R$ 1,7 bi →", "n"),
        ("  dívida bruta financeira ≈ R$ 5,9 bi. Alavancagem 1,1x EBITDA (vs 1,2x no 4T25).", "n"),
        ("• Custo médio da dívida: CDI + 2,09% a.a.", "n"),
        ("• Dívidas vincendas até fim de 2026: R$ 1,2 bi (caixa cobre 1,4x).", "n"),
        ("• Emissões recentes: 13ª (out/25, ≥R$ 1 bi, séries de 5/7/10 anos) e 14ª (mar/26, R$ 1,32 bi).", "n"),
        ("• EBITDA 1T26 R$ 672 mi; geração de caixa operacional R$ 635 mi (95% do EBITDA).", "n"),
        ("", None),
        ("PREMISSAS ESTIMADAS (ajuste com o release/ITR)", "b"),
        ("• Dívida modelada como tranche única consolidada em rolagem (amort. = 0 no default).", "n"),
        ("  Preencha o cronograma real de amortização (ex.: R$ 1,2 bi em 2026) nas células amarelas.", "n"),
        ("• Dívida líquida financeira ampliada (c/ aquisições a pagar + antecipação de recebíveis)", "n"),
        ("  era R$ 5,6 bi no 1T26 — não incluída aqui (só dívida financeira).", "n"),
        ("• Não inclui passivo de arrendamento (IFRS 16), relevante p/ o setor.", "n"),
        ("", None),
        ("FONTES", "b"),
        ("• Release/webinar 1T26 (Nord / Empiricus / XP / Seu Dinheiro); emissões 13ª e 14ª", "s"),
        ("  (Investidor10 / CNN Brasil / ADVFN); SmartFit RI — Debt Issuance.", "s"),
    ],
}

# ----------------------------------------------------------------------------
# Estilos
# ----------------------------------------------------------------------------
F_TITLE = Font(bold=True, size=14, color="1F3864")
F_SUB = Font(italic=True, size=9, color="595959")
F_SECTION = Font(bold=True, size=10, color="FFFFFF")
F_BOLD = Font(bold=True, size=10)
F_NORM = Font(size=10)
F_HDR = Font(bold=True, size=10, color="FFFFFF")
F_NOTE = Font(italic=True, size=8, color="808080")

FILL_SECTION = PatternFill("solid", fgColor="1F3864")
FILL_HDR = PatternFill("solid", fgColor="2E5496")
FILL_INPUT = PatternFill("solid", fgColor="FFF2CC")
FILL_ANCHOR = PatternFill("solid", fgColor="DDEBF7")
FILL_OUT = PatternFill("solid", fgColor="E2EFDA")

THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

NUM = '#,##0.0;(#,##0.0)'
PCT = '0.00%'
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
RIGHT = Alignment(horizontal="right")
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
FONTMAP = {"b": F_BOLD, "n": F_NORM, "s": F_NOTE, "t": F_TITLE}


def build_model(cfg):
    periods = cfg["periods"]
    tranches = cfg["tranches"]
    ncols = len(periods)
    first = 2
    cols = [get_column_letter(first + i) for i in range(ncols)]

    wb = Workbook()

    # ---------------- Aba Leia-me ----------------
    ws0 = wb.active
    ws0.title = "Leia-me"
    ws0.sheet_view.showGridLines = False
    ws0.column_dimensions["A"].width = 2
    ws0.column_dimensions["B"].width = 112
    head = [
        (cfg["title"], "t"), ("", None),
        ("O QUE É", "b"),
        ("Cronograma de dívida bruta, despesa de juros e dívida líquida, montado como motor", "n"),
        ("independente para acoplar ao seu valuation. Trimestral (aberto até 4T27) e anual até 2037.", "n"),
        ("", None),
        ("COMO ACOPLAR (aba 'Modelo_Dívida')", "b"),
        ("1. Amarelo = premissa/input (CDI, spreads, amortizações, captações, dividendos e o FCF livre).", "n"),
        ("2. Ligue a linha 'FCF livre pré-financiamento (LINK)' ao FCF do seu modelo → caixa e dívida", "n"),
        ("   líquida rodam sozinhos.", "n"),
        ("3. Referencie no seu valuation as linhas VERDES da seção de Interface.", "n"),
        ("4. Azul = dado real reportado (âncora), não recalcula.", "n"),
        ("", None),
    ]
    for i, (text, key) in enumerate(head + cfg["readme"] + [
        ("", None),
        ("Gerado por build_debt_model.py — reexecute p/ regenerar. R$ milhões salvo indicado.", "s"),
    ], start=1):
        c = ws0.cell(row=i, column=2, value=text)
        if key:
            c.font = FONTMAP[key]
        c.alignment = LEFT

    # ---------------- Aba Modelo ----------------
    ws = wb.create_sheet("Modelo_Dívida")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 46
    for col in cols:
        ws.column_dimensions[col].width = 11.5
    R = {}

    def sd(cell, fmt=NUM, fill=None, font=F_NORM):
        cell.number_format = fmt
        cell.font = font
        cell.alignment = RIGHT
        cell.border = BORDER
        if fill:
            cell.fill = fill

    def label(row, text, font=F_NORM):
        c = ws.cell(row=row, column=1, value=text)
        c.font = font
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    def section(row, text):
        for i in range(0, ncols + 1):
            ws.cell(row=row, column=1 + i).fill = FILL_SECTION
        c = ws.cell(row=row, column=1, value=text)
        c.font = F_SECTION
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)

    # Cabeçalho
    ws.cell(row=1, column=1, value=cfg["title"]).font = F_TITLE
    ws.cell(row=2, column=1,
            value="R$ milhões  •  Trimestral (aberto até 4T27) e anual até 2037  •  "
                  "amarelo=input, azul=real, verde=interface").font = F_SUB

    HROW = 4
    for col_i, (plabel, _, _, _) in enumerate([("Período", 0, 0, 0)] + list(periods)):
        c = ws.cell(row=HROW, column=1 + col_i, value=plabel)
        c.font = F_HDR
        c.fill = FILL_HDR
        c.alignment = CENTER
        c.border = BORDER

    FROW = 5
    label(FROW, "Fator do período (fração de ano)", F_NOTE)
    for j, (_, _, fator, _) in enumerate(periods):
        sd(ws.cell(row=FROW, column=first + j, value=fator), fmt='0.00', font=F_NOTE)

    # 1. PREMISSAS
    r = 7
    section(r, "1. PREMISSAS (edite as células amarelas)")
    r += 1
    R["cdi"] = r
    label(r, "CDI (% a.a.)")
    for j, (_, _, _, cdi) in enumerate(periods):
        sd(ws.cell(row=r, column=first + j, value=cdi), fmt=PCT,
           fill=(FILL_INPUT if cdi is not None else None))
    spread_cells = []
    for t in tranches:
        r += 1
        label(r, f"Spread {t['name']} (a.a.)")
        sd(ws.cell(row=r, column=first, value=t["spread"]), fmt=PCT, fill=FILL_INPUT)
        spread_cells.append(f"$" + get_column_letter(first) + f"${r}")
    r += 1
    R["ir"] = r
    label(r, "Alíquota IR/CSLL")
    sd(ws.cell(row=r, column=first, value=cfg["ir"]), fmt=PCT, fill=FILL_INPUT)
    CDI = R["cdi"]
    IR = f"$" + get_column_letter(first) + f"${R['ir']}"

    # 2..N Tranches
    trows = []
    for ti, t in enumerate(tranches):
        r += 2
        section(r, f"{2 + ti}. {t['name'].upper()}  ({t['desc']})")
        keys = {}
        for k, lbl, fnt in [
            ("ini", "Saldo inicial", F_NORM), ("capt", "(+) Captações", F_NORM),
            ("amort", "(–) Amortizações", F_NORM), ("fim", "= Saldo final", F_BOLD),
            ("med", "Saldo médio", F_NOTE), ("jur", "Despesa de juros", F_NORM),
        ]:
            r += 1
            keys[k] = r
            label(r, lbl, fnt)
        trows.append(keys)

    cons_num = 2 + len(tranches)
    # Consolidado
    r += 2
    section(r, f"{cons_num}. DÍVIDA CONSOLIDADA")
    r += 1
    R["div_bruta"] = r
    label(r, "Dívida bruta (fim do período)", F_BOLD)
    r += 1
    R["jur_tot"] = r
    label(r, "Despesa de juros total (dívida)", F_BOLD)

    # Caixa
    r += 2
    section(r, f"{cons_num + 1}. CAIXA E DÍVIDA LÍQUIDA")
    for k, lbl, fnt in [
        ("cx_ini", "Caixa inicial", F_NORM),
        ("fcf", "(+) FCF livre pré-financiamento (LINK do seu modelo)", F_NORM),
        ("div_pg", "(–) Dividendos / JCP (input)", F_NORM),
        ("jur_pg", "(–) Juros pagos", F_NOTE),
        ("amort_pg", "(–) Amortizações totais", F_NOTE),
        ("capt_rc", "(+) Captações totais", F_NOTE),
        ("cx_fim", "= Caixa e equivalentes (fim)", F_BOLD),
        ("div_liq", "Dívida líquida", F_BOLD),
    ]:
        r += 1
        R[k] = r
        label(r, lbl, fnt)

    # Interface
    r += 2
    section(r, f"{cons_num + 2}. INTERFACE P/ VALUATION  (referencie estas linhas no seu modelo)")
    for k, lbl, fnt in [
        ("o_bruta", "Dívida bruta (fim)", F_BOLD),
        ("o_caixa", "Caixa e equivalentes (fim)", F_BOLD),
        ("o_liq", "Dívida líquida (fim)", F_BOLD),
        ("o_amort", "Amortizações (saída — financiamento)", F_NORM),
        ("o_capt", "Captações (entrada — financiamento)", F_NORM),
        ("o_jur", "Despesa de juros (P&L, competência)", F_BOLD),
        ("o_jur_ir", "Juros após IR", F_NORM),
        ("o_shield", "Economia fiscal dos juros (tax shield)", F_NORM),
        ("o_ebitda", "EBITDA LTM (input opcional — link do seu modelo)", F_NORM),
        ("o_lev", "Dívida líquida / EBITDA (x)", F_NORM),
    ]:
        r += 1
        R[k] = r
        label(r, lbl, fnt)
    LASTROW = r

    # ---------------- Preenchimento ----------------
    for j, (plabel, kind, fator, cdi) in enumerate(periods):
        L = cols[j]
        prev = cols[j - 1] if j > 0 else None
        is_anchor = kind == "anchor"
        is_real = kind == "real"
        cash_locked = kind in ("anchor", "real")

        def put(row, value, fmt=NUM, fill=None, font=F_NORM):
            sd(ws.cell(row=row, column=first + j, value=value), fmt=fmt, fill=fill, font=font)

        # tranches
        for ti, (t, tr) in enumerate(zip(tranches, trows)):
            if is_anchor:
                put(tr["fim"], t["opening"], fill=FILL_ANCHOR, font=F_BOLD)
            else:
                put(tr["ini"], f"={prev}{tr['fim']}")
                put(tr["capt"], 0, fill=FILL_INPUT)
                put(tr["amort"], t["amort"].get(plabel, 0), fill=FILL_INPUT)
                put(tr["fim"], f"={L}{tr['ini']}+{L}{tr['capt']}-{L}{tr['amort']}", font=F_BOLD)
                put(tr["med"], f"=({L}{tr['ini']}+{L}{tr['fim']})/2", font=F_NOTE)
                put(tr["jur"], f"={L}{tr['med']}*({L}{CDI}+{spread_cells[ti]})*{L}{FROW}")

        # consolidado
        fim_sum = "+".join(f"{L}{tr['fim']}" for tr in trows)
        put(R["div_bruta"], f"={fim_sum}", font=F_BOLD,
            fill=(FILL_ANCHOR if is_anchor else None))
        if not is_anchor:
            jur_sum = "+".join(f"{L}{tr['jur']}" for tr in trows)
            put(R["jur_tot"], f"={jur_sum}", font=F_BOLD)

        # caixa
        amort_sum = "+".join(f"{L}{tr['amort']}" for tr in trows)
        capt_sum = "+".join(f"{L}{tr['capt']}" for tr in trows)
        if cash_locked:
            if not is_anchor:
                put(R["jur_pg"], f"={L}{R['jur_tot']}", font=F_NOTE)
                put(R["amort_pg"], f"={amort_sum}", font=F_NOTE)
                put(R["capt_rc"], f"={capt_sum}", font=F_NOTE)
            put(R["cx_fim"], cfg["cash_anchors"][plabel], fill=FILL_ANCHOR, font=F_BOLD)
        else:
            put(R["cx_ini"], f"={prev}{R['cx_fim']}")
            put(R["fcf"], 0, fill=FILL_INPUT)
            put(R["div_pg"], 0, fill=FILL_INPUT)
            put(R["jur_pg"], f"={L}{R['jur_tot']}", font=F_NOTE)
            put(R["amort_pg"], f"={amort_sum}", font=F_NOTE)
            put(R["capt_rc"], f"={capt_sum}", font=F_NOTE)
            put(R["cx_fim"],
                f"={L}{R['cx_ini']}+{L}{R['fcf']}-{L}{R['div_pg']}-{L}{R['jur_pg']}"
                f"-{L}{R['amort_pg']}+{L}{R['capt_rc']}", font=F_BOLD)

        put(R["div_liq"], f"={L}{R['div_bruta']}-{L}{R['cx_fim']}", font=F_BOLD,
            fill=(FILL_ANCHOR if cash_locked else None))

        # interface
        put(R["o_bruta"], f"={L}{R['div_bruta']}", fill=FILL_OUT, font=F_BOLD)
        put(R["o_caixa"], f"={L}{R['cx_fim']}", fill=FILL_OUT, font=F_BOLD)
        put(R["o_liq"], f"={L}{R['div_liq']}", fill=FILL_OUT, font=F_BOLD)
        if not is_anchor:
            put(R["o_amort"], f"={amort_sum}", fill=FILL_OUT)
            put(R["o_capt"], f"={capt_sum}", fill=FILL_OUT)
            put(R["o_jur"], f"={L}{R['jur_tot']}", fill=FILL_OUT, font=F_BOLD)
            put(R["o_jur_ir"], f"={L}{R['jur_tot']}*(1-{IR})", fill=FILL_OUT)
            put(R["o_shield"], f"={L}{R['jur_tot']}*{IR}", fill=FILL_OUT)
        put(R["o_ebitda"], None, fill=FILL_INPUT)
        put(R["o_lev"], f"=IFERROR({L}{R['o_liq']}/{L}{R['o_ebitda']},\"\")",
            fmt='0.00"x"', fill=FILL_OUT)

    note_row = LASTROW + 2
    nc = ws.cell(row=note_row, column=1,
                 value="Nota: juros em regime de competência (proxy p/ caixa). Ligue a linha de FCF "
                       "livre ao seu modelo para o caixa e a dívida líquida rodarem sozinhos. "
                       "Preencha as amortizações/captações reais nas células amarelas.")
    nc.font = F_NOTE
    nc.alignment = LEFT
    ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=ncols + 1)

    ws.freeze_panes = "B5"
    wb.save(cfg["file"])
    print(f"OK -> {cfg['file']}  ({ncols} períodos, {LASTROW} linhas)")


if __name__ == "__main__":
    build_model(VIVARA)
    build_model(SMARTFIT)
