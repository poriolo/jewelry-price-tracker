"""
Gera o modelo de dívida da Vivara (VIVA3) em Excel — módulo acoplável a um
valuation. Periodicidade: trimestral (4T25 real até 4T27E) e depois anual
(2028E–2030E). Todas as células são fórmulas vivas; premissas ficam em amarelo,
âncoras reais em azul, saídas de acoplamento em verde.

Rode:  python3 build_debt_model.py
Saída: vivara_debt_model.xlsx

Fontes dos dados-âncora estão na aba "Leia-me" e no README.md desta pasta.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ----------------------------------------------------------------------------
# Definição de períodos (colunas B..M)
# (label, tipo, fator_ano, cdi_aa)
#   tipo: "anchor" = 4T25 (só saldos reais), "real" = 1T26 (saldos reais +
#         juros calculados, caixa travado), "proj" = projeção completa
# ----------------------------------------------------------------------------
PERIODS = [
    ("4T25\n(Real)", "anchor", None,  None),
    ("1T26\n(Real)", "real",   0.25,  0.1490),
    ("2T26E",        "proj",   0.25,  0.1490),
    ("3T26E",        "proj",   0.25,  0.1475),
    ("4T26E",        "proj",   0.25,  0.1450),
    ("1T27E",        "proj",   0.25,  0.1400),
    ("2T27E",        "proj",   0.25,  0.1350),
    ("3T27E",        "proj",   0.25,  0.1300),
    ("4T27E",        "proj",   0.25,  0.1250),
    ("2028E",        "proj",   1.00,  0.1100),
    ("2029E",        "proj",   1.00,  0.1050),
    ("2030E",        "proj",   1.00,  0.1000),
]
FIRST_COL = 2  # coluna B
COLS = [get_column_letter(FIRST_COL + i) for i in range(len(PERIODS))]

# Âncoras reais (R$ milhões) — ver "Leia-me"
DEB_INICIAL_4T25   = 300.0    # Debêntures VIVA11 (1a emissão)
DIVIDA_BRUTA_4T25  = 531.3    # dívida bruta 31/12/2025
CAIXA_4T25         = 398.6    # caixa + equivalentes 31/12/2025
CAIXA_1T26         = 284.7    # estimado p/ fechar DL reportada de 246,6 (531,3-246,6)

# Amortização das debêntures: 50% em 27/08/2029 e 50% em 27/08/2030
DEB_AMORT = {"2029E": 150.0, "2030E": 150.0}

# Premissas escalares
SPREAD_DEB    = 0.0070   # 100% CDI + 0,70% a.a.
SPREAD_OUTRAS = 0.0200   # premissa (bancos/FX/capital de giro)
ALIQUOTA_IR   = 0.34

# ----------------------------------------------------------------------------
# Estilos
# ----------------------------------------------------------------------------
F_TITLE   = Font(bold=True, size=14, color="1F3864")
F_SUB     = Font(italic=True, size=9, color="595959")
F_SECTION = Font(bold=True, size=10, color="FFFFFF")
F_BOLD    = Font(bold=True, size=10)
F_NORM    = Font(size=10)
F_HDR     = Font(bold=True, size=10, color="FFFFFF")
F_NOTE    = Font(italic=True, size=8, color="808080")

FILL_SECTION = PatternFill("solid", fgColor="1F3864")
FILL_HDR     = PatternFill("solid", fgColor="2E5496")
FILL_INPUT   = PatternFill("solid", fgColor="FFF2CC")  # amarelo = editável
FILL_ANCHOR  = PatternFill("solid", fgColor="DDEBF7")  # azul = dado real
FILL_OUT     = PatternFill("solid", fgColor="E2EFDA")  # verde = interface
FILL_TOTAL   = PatternFill("solid", fgColor="F2F2F2")

THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

NUM = '#,##0.0;(#,##0.0)'
PCT = '0.00%'

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
RIGHT  = Alignment(horizontal="right")
LEFT   = Alignment(horizontal="left", vertical="center", wrap_text=True)


def style_data(cell, fmt=NUM, fill=None, font=F_NORM):
    cell.number_format = fmt
    cell.font = font
    cell.alignment = RIGHT
    cell.border = BORDER
    if fill:
        cell.fill = fill


# ----------------------------------------------------------------------------
# Workbook
# ----------------------------------------------------------------------------
wb = Workbook()

# ============================ Aba: Leia-me ==================================
ws0 = wb.active
ws0.title = "Leia-me"
ws0.sheet_view.showGridLines = False
ws0.column_dimensions["A"].width = 2
ws0.column_dimensions["B"].width = 110

readme = [
    ("VIVARA (VIVA3) — Fluxo de Dívida | Módulo acoplável ao valuation", F_TITLE),
    ("", None),
    ("O QUE É", F_BOLD),
    ("Cronograma de dívida bruta, despesa de juros e dívida líquida da Vivara, montado", F_NORM),
    ("como motor independente para você acoplar ao seu modelo de valuation. Periodicidade", F_NORM),
    ("trimestral do 4T25 (real) até o 4T27E e depois anual (2028E–2030E).", F_NORM),
    ("", None),
    ("COMO ACOPLAR (aba 'Modelo_Dívida')", F_BOLD),
    ("1. Células AMARELAS são premissas/inputs — edite à vontade (CDI, spreads, amortizações,", F_NORM),
    ("   captações, dividendos e o FCF livre que vem do SEU modelo).", F_NORM),
    ("2. Ligue a linha 'Fluxo de caixa livre pré-financiamento (LINK)' ao FCF do seu modelo;", F_NORM),
    ("   o caixa e a dívida líquida passam a rodar sozinhos.", F_NORM),
    ("3. Referencie no seu valuation as linhas da seção 6 (VERDES): dívida bruta, dívida líquida,", F_NORM),
    ("   despesa de juros, juros após IR e escudo fiscal.", F_NORM),
    ("4. Células AZUIS são dados reais reportados (âncoras) — não recalculam.", F_NORM),
    ("", None),
    ("PREMISSAS-CHAVE (reais, pesquisadas)", F_BOLD),
    ("• Debênture VIVA11 (1ª emissão): R$ 300 mi, série única, 27/08/2025→27/08/2030,", F_NORM),
    ("  100% do CDI + 0,70% a.a., juros semestrais; amortização 50% em 2029 e 50% em 2030;", F_NORM),
    ("  recursos usados p/ liquidar CCBs (~R$ 102 mi) — liability management.", F_NORM),
    ("• 31/12/2025: dívida bruta R$ 531,3 mi | caixa R$ 398,6 mi | dívida líquida R$ 132,6 mi.", F_NORM),
    ("• 1T26: dívida líquida R$ 246,6 mi | alavancagem 0,3x ND/EBITDA.", F_NORM),
    ("", None),
    ("PREMISSAS ESTIMADAS (ajuste com o ITR quando tiver)", F_BOLD),
    ("• 'Outras dívidas' = dívida bruta 4T25 − debêntures = R$ 231,3 mi (bancos/FX/capital de giro).", F_NORM),
    ("  Modeladas como rolagem (amortização default = 0); preencha o cronograma real na aba.", F_NORM),
    ("• Spread 'outras dívidas' = CDI + 2,00% a.a. (premissa).", F_NORM),
    ("• Curva de CDI: ~14,9% (2026) caindo p/ ~10% (2030) — edite na linha CDI.", F_NORM),
    ("• Caixa 1T26 (R$ 284,7 mi) é derivado p/ fechar a DL reportada de R$ 246,6 mi.", F_NORM),
    ("• 'Despesa de juros' aqui é só a da dívida financeira; o resultado financeiro reportado", F_NORM),
    ("  (−R$ 15 mi no 1T26) inclui derivativos/MtM e não bate 1:1 com esta linha.", F_NORM),
    ("", None),
    ("FONTES", F_BOLD),
    ("• Release/notícias de resultados 1T26 (Money Times, Nord, Análise de Ações).", F_NOTE),
    ("• Escritura e características VIVA11 (debentures.com.br, Mais Retorno, ANBIMA).", F_NOTE),
    ("• Emissão R$ 300 mi p/ liquidar CCBs (Lex Legal — Lefosse/Mattos Filho).", F_NOTE),
    ("• Dados de dívida bruta/caixa 31/12/2025 (DF 2025 Vivara Participações).", F_NOTE),
    ("", None),
    ("Gerado por build_debt_model.py — reexecute p/ regenerar. R$ milhões salvo indicado.", F_SUB),
]
for i, (text, font) in enumerate(readme, start=1):
    c = ws0.cell(row=i, column=2, value=text)
    if font:
        c.font = font
    c.alignment = LEFT

# ============================ Aba: Modelo ===================================
ws = wb.create_sheet("Modelo_Dívida")
ws.sheet_view.showGridLines = False
ws.column_dimensions["A"].width = 46
for col in COLS:
    ws.column_dimensions[col].width = 11.5

R = {}  # mapa nome->linha para as fórmulas


def label(row, text, font=F_NORM, fill=None, indent=0):
    c = ws.cell(row=row, column=1, value=text)
    c.font = font
    c.alignment = Alignment(horizontal="left", vertical="center", indent=indent, wrap_text=True)
    if fill:
        c.fill = fill
    return c


def section(row, text):
    for i in range(0, len(PERIODS) + 1):
        cc = ws.cell(row=row, column=1 + i)
        cc.fill = FILL_SECTION
    c = ws.cell(row=row, column=1, value=text)
    c.font = F_SECTION
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)


# --- Cabeçalho ---
ws.cell(row=1, column=1, value="VIVARA (VIVA3) — Fluxo de Dívida (módulo de valuation)").font = F_TITLE
ws.cell(row=2, column=1,
        value="R$ milhões, salvo indicado  •  Trimestral 4T25–4T27, anual 2028–2030  •  amarelo=input, azul=real, verde=interface").font = F_SUB

# --- Linha de períodos (row 4) ---
HROW = 4
hc = ws.cell(row=HROW, column=1, value="Período")
hc.font = F_HDR
hc.fill = FILL_HDR
hc.alignment = CENTER
hc.border = BORDER
for j, (plabel, _, _, _) in enumerate(PERIODS):
    c = ws.cell(row=HROW, column=FIRST_COL + j, value=plabel)
    c.font = F_HDR
    c.fill = FILL_HDR
    c.alignment = CENTER
    c.border = BORDER

# --- Fator do período (row 5) ---
FROW = 5
label(FROW, "Fator do período (fração de ano)", F_NOTE)
for j, (_, _, fator, _) in enumerate(PERIODS):
    c = ws.cell(row=FROW, column=FIRST_COL + j, value=fator)
    style_data(c, fmt='0.00', font=F_NOTE)

# ---------------------------------------------------------------------------
# 1. PREMISSAS
# ---------------------------------------------------------------------------
r = 7
section(r, "1. PREMISSAS (edite as células amarelas)")
r += 1
R["cdi"] = r
label(r, "CDI (% a.a.)", F_NORM)
for j, (_, _, _, cdi) in enumerate(PERIODS):
    c = ws.cell(row=r, column=FIRST_COL + j, value=cdi)
    style_data(c, fmt=PCT, fill=(FILL_INPUT if cdi is not None else None))
r += 1
R["spread_deb"] = r
label(r, "Spread debêntures VIVA11 (a.a.)", F_NORM)
c = ws.cell(row=r, column=FIRST_COL, value=SPREAD_DEB)
style_data(c, fmt=PCT, fill=FILL_INPUT)
r += 1
R["spread_out"] = r
label(r, "Spread outras dívidas (a.a.)", F_NORM)
c = ws.cell(row=r, column=FIRST_COL, value=SPREAD_OUTRAS)
style_data(c, fmt=PCT, fill=FILL_INPUT)
r += 1
R["ir"] = r
label(r, "Alíquota IR/CSLL", F_NORM)
c = ws.cell(row=r, column=FIRST_COL, value=ALIQUOTA_IR)
style_data(c, fmt=PCT, fill=FILL_INPUT)

CDI = R["cdi"]
SDEB = f"$" + get_column_letter(FIRST_COL) + f"${R['spread_deb']}"
SOUT = f"$" + get_column_letter(FIRST_COL) + f"${R['spread_out']}"
IR = f"$" + get_column_letter(FIRST_COL) + f"${R['ir']}"

# ---------------------------------------------------------------------------
# 2. DEBÊNTURES VIVA11
# ---------------------------------------------------------------------------
r += 2
section(r, "2. DEBÊNTURES VIVA11  (100% CDI + 0,70% • amort. 50% 2029 / 50% 2030)")
r += 1
R["deb_ini"] = r; label(r, "Saldo inicial", F_NORM)
r += 1
R["deb_capt"] = r; label(r, "(+) Captações", F_NORM)
r += 1
R["deb_amort"] = r; label(r, "(–) Amortizações", F_NORM)
r += 1
R["deb_fim"] = r; label(r, "= Saldo final", F_BOLD)
r += 1
R["deb_med"] = r; label(r, "Saldo médio", F_NOTE)
r += 1
R["deb_jur"] = r; label(r, "Despesa de juros", F_NORM)

# ---------------------------------------------------------------------------
# 3. OUTRAS DÍVIDAS
# ---------------------------------------------------------------------------
r += 2
section(r, "3. OUTRAS DÍVIDAS  (bancos / FX / capital de giro — premissa: rolagem)")
r += 1
R["out_ini"] = r; label(r, "Saldo inicial", F_NORM)
r += 1
R["out_capt"] = r; label(r, "(+) Captações (input)", F_NORM)
r += 1
R["out_amort"] = r; label(r, "(–) Amortizações (input)", F_NORM)
r += 1
R["out_fim"] = r; label(r, "= Saldo final", F_BOLD)
r += 1
R["out_med"] = r; label(r, "Saldo médio", F_NOTE)
r += 1
R["out_jur"] = r; label(r, "Despesa de juros", F_NORM)

# ---------------------------------------------------------------------------
# 4. DÍVIDA CONSOLIDADA
# ---------------------------------------------------------------------------
r += 2
section(r, "4. DÍVIDA CONSOLIDADA")
r += 1
R["div_bruta"] = r; label(r, "Dívida bruta (fim do período)", F_BOLD)
r += 1
R["jur_tot"] = r; label(r, "Despesa de juros total (dívida)", F_BOLD)

# ---------------------------------------------------------------------------
# 5. CAIXA E DÍVIDA LÍQUIDA
# ---------------------------------------------------------------------------
r += 2
section(r, "5. CAIXA E DÍVIDA LÍQUIDA")
r += 1
R["cx_ini"] = r; label(r, "Caixa inicial", F_NORM)
r += 1
R["fcf"] = r; label(r, "(+) FCF livre pré-financiamento (LINK do seu modelo)", F_NORM)
r += 1
R["div_pg"] = r; label(r, "(–) Dividendos / JCP (input)", F_NORM)
r += 1
R["jur_pg"] = r; label(r, "(–) Juros pagos", F_NOTE)
r += 1
R["amort_pg"] = r; label(r, "(–) Amortizações totais", F_NOTE)
r += 1
R["capt_rc"] = r; label(r, "(+) Captações totais", F_NOTE)
r += 1
R["cx_fim"] = r; label(r, "= Caixa e equivalentes (fim)", F_BOLD)
r += 1
R["div_liq"] = r; label(r, "Dívida líquida", F_BOLD)

# ---------------------------------------------------------------------------
# 6. INTERFACE P/ VALUATION
# ---------------------------------------------------------------------------
r += 2
section(r, "6. INTERFACE P/ VALUATION  (referencie estas linhas no seu modelo)")
r += 1
R["o_bruta"] = r; label(r, "Dívida bruta (fim)", F_BOLD)
r += 1
R["o_caixa"] = r; label(r, "Caixa e equivalentes (fim)", F_BOLD)
r += 1
R["o_liq"] = r; label(r, "Dívida líquida (fim)", F_BOLD)
r += 1
R["o_amort"] = r; label(r, "Amortizações (saída — financiamento)", F_NORM)
r += 1
R["o_capt"] = r; label(r, "Captações (entrada — financiamento)", F_NORM)
r += 1
R["o_jur"] = r; label(r, "Despesa de juros (P&L, competência)", F_BOLD)
r += 1
R["o_jur_ir"] = r; label(r, "Juros após IR", F_NORM)
r += 1
R["o_shield"] = r; label(r, "Economia fiscal dos juros (tax shield)", F_NORM)
r += 1
R["o_ebitda"] = r; label(r, "EBITDA LTM (input opcional — link do seu modelo)", F_NORM)
r += 1
R["o_lev"] = r; label(r, "Dívida líquida / EBITDA (x)", F_NORM)
LASTROW = r

# ---------------------------------------------------------------------------
# Preenchimento coluna a coluna
# ---------------------------------------------------------------------------
for j, (plabel, kind, fator, cdi) in enumerate(PERIODS):
    L = COLS[j]
    prev = COLS[j - 1] if j > 0 else None
    is_anchor = kind == "anchor"       # 4T25
    is_real = kind == "real"           # 1T26
    cash_locked = kind in ("anchor", "real")

    def put(key, value, fmt=NUM, fill=None, font=F_NORM):
        c = ws.cell(row=R[key], column=FIRST_COL + j, value=value)
        style_data(c, fmt=fmt, fill=fill, font=font)
        return c

    # ---- DEBÊNTURES ----
    if is_anchor:
        put("deb_fim", DEB_INICIAL_4T25, fill=FILL_ANCHOR, font=F_BOLD)
    else:
        put("deb_ini", f"={prev}{R['deb_fim']}")
        put("deb_capt", 0, fill=FILL_INPUT)
        amort = DEB_AMORT.get(plabel, 0)
        put("deb_amort", amort, fill=FILL_INPUT)
        put("deb_fim", f"={L}{R['deb_ini']}+{L}{R['deb_capt']}-{L}{R['deb_amort']}", font=F_BOLD)
        put("deb_med", f"=({L}{R['deb_ini']}+{L}{R['deb_fim']})/2", font=F_NOTE)
        put("deb_jur", f"={L}{R['deb_med']}*({L}{CDI}+{SDEB})*{L}{FROW}")

    # ---- OUTRAS DÍVIDAS ----
    if is_anchor:
        put("out_fim", f"={L}{R['div_bruta']}-{L}{R['deb_fim']}", fill=FILL_ANCHOR, font=F_BOLD)
    else:
        put("out_ini", f"={prev}{R['out_fim']}")
        put("out_capt", 0, fill=FILL_INPUT)
        put("out_amort", 0, fill=FILL_INPUT)
        put("out_fim", f"={L}{R['out_ini']}+{L}{R['out_capt']}-{L}{R['out_amort']}", font=F_BOLD)
        put("out_med", f"=({L}{R['out_ini']}+{L}{R['out_fim']})/2", font=F_NOTE)
        put("out_jur", f"={L}{R['out_med']}*({L}{CDI}+{SOUT})*{L}{FROW}")

    # ---- CONSOLIDADO ----
    if is_anchor:
        put("div_bruta", DIVIDA_BRUTA_4T25, fill=FILL_ANCHOR, font=F_BOLD)
    else:
        put("div_bruta", f"={L}{R['deb_fim']}+{L}{R['out_fim']}", font=F_BOLD)
        put("jur_tot", f"={L}{R['deb_jur']}+{L}{R['out_jur']}", font=F_BOLD)

    # ---- CAIXA & DÍVIDA LÍQUIDA ----
    if is_anchor:
        put("cx_fim", CAIXA_4T25, fill=FILL_ANCHOR, font=F_BOLD)
    elif is_real:
        # 1T26: caixa travado (âncora), sem roll; mostra juros calculados
        put("jur_pg", f"={L}{R['jur_tot']}", font=F_NOTE)
        put("amort_pg", f"={L}{R['deb_amort']}+{L}{R['out_amort']}", font=F_NOTE)
        put("capt_rc", f"={L}{R['deb_capt']}+{L}{R['out_capt']}", font=F_NOTE)
        put("cx_fim", CAIXA_1T26, fill=FILL_ANCHOR, font=F_BOLD)
    else:
        put("cx_ini", f"={prev}{R['cx_fim']}")
        put("fcf", 0, fill=FILL_INPUT)
        put("div_pg", 0, fill=FILL_INPUT)
        put("jur_pg", f"={L}{R['jur_tot']}", font=F_NOTE)
        put("amort_pg", f"={L}{R['deb_amort']}+{L}{R['out_amort']}", font=F_NOTE)
        put("capt_rc", f"={L}{R['deb_capt']}+{L}{R['out_capt']}", font=F_NOTE)
        put("cx_fim",
            f"={L}{R['cx_ini']}+{L}{R['fcf']}-{L}{R['div_pg']}-{L}{R['jur_pg']}"
            f"-{L}{R['amort_pg']}+{L}{R['capt_rc']}", font=F_BOLD)

    put("div_liq", f"={L}{R['div_bruta']}-{L}{R['cx_fim']}", font=F_BOLD,
        fill=(FILL_ANCHOR if cash_locked else None))

    # ---- INTERFACE (verde) ----
    put("o_bruta", f"={L}{R['div_bruta']}", fill=FILL_OUT, font=F_BOLD)
    put("o_caixa", f"={L}{R['cx_fim']}", fill=FILL_OUT, font=F_BOLD)
    put("o_liq", f"={L}{R['div_liq']}", fill=FILL_OUT, font=F_BOLD)
    if not is_anchor:
        put("o_amort", f"={L}{R['deb_amort']}+{L}{R['out_amort']}", fill=FILL_OUT)
        put("o_capt", f"={L}{R['deb_capt']}+{L}{R['out_capt']}", fill=FILL_OUT)
        put("o_jur", f"={L}{R['jur_tot']}", fill=FILL_OUT, font=F_BOLD)
        put("o_jur_ir", f"={L}{R['jur_tot']}*(1-{IR})", fill=FILL_OUT)
        put("o_shield", f"={L}{R['jur_tot']}*{IR}", fill=FILL_OUT)
    put("o_ebitda", None, fmt=NUM, fill=FILL_INPUT)
    put("o_lev", f"=IFERROR({L}{R['o_liq']}/{L}{R['o_ebitda']},\"\")", fmt='0.00"x"', fill=FILL_OUT)

# Nota de rodapé
note_row = LASTROW + 2
nc = ws.cell(row=note_row, column=1,
             value="Nota: juros em regime de competência (proxy p/ caixa). Debêntures pagam juros "
                   "semestrais (fev/ago) e amortizam 50% em 2029 e 50% em 2030. Ligue a linha de "
                   "FCF livre ao seu modelo para o caixa e a dívida líquida rodarem sozinhos.")
nc.font = F_NOTE
nc.alignment = LEFT
ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=len(PERIODS) + 1)

ws.freeze_panes = "B5"

out = "vivara_debt_model.xlsx"
wb.save(out)
print(f"OK -> {out}  ({len(PERIODS)} períodos, {LASTROW} linhas)")
